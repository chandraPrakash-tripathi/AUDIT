#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Excel Handler for GST Reconciliation System

This module provides functions for reading from and writing to Excel files,
which are used extensively throughout the GST reconciliation process.
"""

import os
import pandas as pd
import numpy as np
import logging
from datetime import datetime
import xlrd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart

# Set up logging
logging.basicConfig(level=logging.INFO, 
                    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def read_excel_file(file_path, sheet_name=None):
    """
    Read data from Excel file.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str or list, optional): Sheet name(s) to read. 
                                           If None, reads all sheets.
    
    Returns:
        pandas.DataFrame or dict of DataFrames: Data from the Excel file
    """
    try:
        logger.info(f"Reading Excel file: {file_path}")
        
        if sheet_name is None:
            # Get all sheet names first to inspect what's available
            xlsx = pd.ExcelFile(file_path)
            available_sheets = xlsx.sheet_names
            logger.info(f"Available sheets: {available_sheets}")
            
            # Read all sheets into a dictionary of DataFrames
            data = {}
            for sheet in available_sheets:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet)
                    if not df.empty:
                        data[sheet] = df
                except Exception as e:
                    logger.warning(f"Could not read sheet '{sheet}': {str(e)}")
            
            # If only one sheet was read, return just the DataFrame instead of a dict
            if len(data) == 1:
                return list(data.values())[0]
            return data
            
        else:
            # Read specific sheet(s)
            return pd.read_excel(file_path, sheet_name=sheet_name)
            
    except Exception as e:
        logger.error(f"Error reading Excel file: {str(e)}")
        raise IOError(f"Failed to read Excel file: {str(e)}")

def write_excel_file(data, file_path, sheet_name="Sheet1", include_index=False):
    """
    Write data to Excel file.
    
    Args:
        data (pandas.DataFrame): Data to write
        file_path (str): Path to save the Excel file
        sheet_name (str, optional): Sheet name. Defaults to "Sheet1".
        include_index (bool, optional): Whether to include the DataFrame index.
                                       Defaults to False.
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        logger.info(f"Writing data to Excel file: {file_path}")
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        
        # Write to Excel
        data.to_excel(file_path, sheet_name=sheet_name, index=include_index)
        
        logger.info(f"Successfully wrote data to {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error writing to Excel file: {str(e)}")
        raise IOError(f"Failed to write to Excel file: {str(e)}")

def save_excel_file(data, file_path, styling=True):
    """
    Save data to an Excel file with optional styling.
    
    Args:
        data (pandas.DataFrame or dict): DataFrame or dict of DataFrames to save
        file_path (str): Path to save the Excel file
        styling (bool, optional): Whether to apply styling. Defaults to True.
    
    Returns:
        bool: True if successful, False otherwise
    """
    try:
        logger.info(f"Saving data to Excel file: {file_path}")
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
        
        if isinstance(data, dict):
            # Multiple sheets
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                for sheet_name, df in data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    if styling:
                        _apply_styling(writer, df, sheet_name)
        else:
            # Single DataFrame
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                data.to_excel(writer, index=False)
                
                if styling:
                    _apply_styling(writer, data)
        
        logger.info(f"Successfully saved data to {file_path}")
        return True
        
    except Exception as e:
        logger.error(f"Error saving Excel file: {str(e)}")
        raise IOError(f"Failed to save Excel file: {str(e)}")

def _apply_styling(writer, df, sheet_name="Sheet1"):
    """
    Apply styling to an Excel worksheet.
    
    Args:
        writer (pandas.ExcelWriter): Excel writer object
        df (pandas.DataFrame): Data being written
        sheet_name (str, optional): Sheet name. Defaults to "Sheet1".
    """
    workbook = writer.book
    try:
        worksheet = writer.sheets[sheet_name]
    except:
        worksheet = workbook.active
    
    # Define styles
    header_fill = PatternFill(start_color="D7E4BC", end_color="D7E4BC", fill_type="solid")
    header_font = Font(bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border
    
    # Style data cells
    for row_idx in range(2, len(df) + 2):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = border
            
            # Align numbers to right
            if isinstance(df.iloc[row_idx-2, col_idx-1], (int, float)):
                cell.alignment = Alignment(horizontal='right')
    
    # Auto-adjust column widths
    for col_idx, column in enumerate(df.columns, 1):
        column_width = max(
            len(str(column)),  # Column header width
            df[column].astype(str).str.len().max(),  # Max data width
            10  # Minimum width
        ) + 2  # Add padding
        
        col_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[col_letter].width = min(column_width, 40)  # Cap at 40
    
    # Add alternating row colors for readability if more than 5 rows
    if len(df) > 5:
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_idx in range(3, len(df) + 2, 2):  # Start from 3 (2nd data row) and do every other row
            for col_idx in range(1, len(df.columns) + 1):
                worksheet.cell(row=row_idx, column=col_idx).fill = light_fill

def create_workbook_with_charts(data, charts_config=None):
    """
    Create an Excel workbook with charts for visualization.
    
    Args:
        data (dict): Dictionary with sheet_name -> DataFrame mapping
        charts_config (list, optional): Configuration for charts
    
    Returns:
        openpyxl.Workbook: Workbook with data and charts
    """
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Add data sheets
    for sheet_name, df in data.items():
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
        
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Add charts if configured
    if charts_config:
        if 'Charts' not in wb.sheetnames:
            charts_sheet = wb.create_sheet(title='Charts')
        else:
            charts_sheet = wb['Charts']
        
        current_row = 1
        
        for chart_config in charts_config:
            chart_type = chart_config.get('type', 'bar')
            source_sheet = chart_config.get('sheet')
            title = chart_config.get('title', f'Chart from {source_sheet}')
            categories_col = chart_config.get('categories_col', 1)
            values_col = chart_config.get('values_col', 2)
            start_row = chart_config.get('start_row', 2)
            end_row = chart_config.get('end_row', None)
            
            if source_sheet not in wb.sheetnames:
                logger.warning(f"Sheet {source_sheet} not found for chart")
                continue
                
            ws = wb[source_sheet]
            
            # Get end row if not specified
            if end_row is None:
                end_row = ws.max_row
            
            # Create chart
            if chart_type.lower() == 'bar':
                chart = BarChart()
            elif chart_type.lower() == 'pie':
                chart = PieChart()
            else:
                logger.warning(f"Unsupported chart type: {chart_type}")
                continue
            
            chart.title = title
            
            # Set categories (x-axis)
            cats = Reference(ws, min_col=categories_col, min_row=start_row, 
                           max_row=end_row)
            chart.set_categories(cats)
            
            # Set values
            values = Reference(ws, min_col=values_col, min_row=start_row-1,  # Include header
                             max_row=end_row, max_col=values_col)
            chart.add_data(values, titles_from_data=True)
            
            # Add to charts sheet
            charts_sheet.add_chart(chart, f"A{current_row}")
            current_row += 15  # Move down for next chart
    
    return wb

def load_excel_file(file_path, sheet_name=None):
    """
    Load Excel file for viewing or processing.
    
    Args:
        file_path (str): Path to the Excel file
        sheet_name (str or list, optional): Sheet name(s) to read.
                                          If None, reads the first sheet.
    
    Returns:
        pandas.DataFrame: Data from the Excel file
    """
    try:
        logger.info(f"Loading Excel file for viewing: {file_path}")
        
        if sheet_name is None:
            # If no sheet specified, read the first sheet
            df = pd.read_excel(file_path, sheet_name=0)
        else:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
        return df
        
    except Exception as e:
        logger.error(f"Error loading Excel file: {str(e)}")
        raise IOError(f"Failed to load Excel file: {str(e)}")

def get_sheet_names(file_path):
    """
    Get all sheet names from an Excel file.
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        list: List of sheet names
    """
    try:
        xlsx = pd.ExcelFile(file_path)
        return xlsx.sheet_names
    except Exception as e:
        logger.error(f"Error getting sheet names: {str(e)}")
        return []

def combine_excel_files(file_paths, sheet_name=None):
    """
    Combine data from multiple Excel files.
    
    Args:
        file_paths (list): List of paths to Excel files
        sheet_name (str, optional): Specific sheet to read from each file
                                   If None, reads the first sheet.
    
    Returns:
        pandas.DataFrame: Combined data
    """
    combined_df = pd.DataFrame()
    
    for file_path in file_paths:
        try:
            if sheet_name is None:
                df = pd.read_excel(file_path)
            else:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
            # Add source file info
            df['Source_File'] = os.path.basename(file_path)
            
            # Append to combined DataFrame
            combined_df = pd.concat([combined_df, df], ignore_index=True)
            
        except Exception as e:
            logger.warning(f"Could not read {file_path}: {str(e)}")
    
    return combined_df

def detect_excel_structure(file_path):
    """
    Detect the structure of an Excel file (headers, data types).
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        dict: Information about file structure
    """
    try:
        # Read file without specifying dtypes to auto-detect
        xlsx = pd.ExcelFile(file_path)
        sheet_names = xlsx.sheet_names
        
        structure = {
            'filename': os.path.basename(file_path),
            'sheets': [],
            'last_modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
        }
        
        for sheet in sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet, nrows=100)  # Read first 100 rows
            
            sheet_info = {
                'name': sheet,
                'row_count': len(pd.read_excel(file_path, sheet_name=sheet)),
                'column_count': len(df.columns),
                'columns': []
            }
            
            for col_name in df.columns:
                col_info = {
                    'name': col_name,
                    'dtype': str(df[col_name].dtype),
                    'has_nulls': df[col_name].isnull().any(),
                    'unique_values': df[col_name].nunique(),
                    'sample': df[col_name].iloc[0] if not df.empty else None
                }
                sheet_info['columns'].append(col_info)
            
            structure['sheets'].append(sheet_info)
        
        return structure
        
    except Exception as e:
        logger.error(f"Error detecting Excel structure: {str(e)}")
        return {'error': str(e)}